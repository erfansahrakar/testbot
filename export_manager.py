"""
✅ FEATURE #4: Export Manager (نسخه Fix شده)
دانلود گزارشات به صورت Excel/CSV
"""
import logging
import os
import tempfile
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from config import ADMIN_ID

logger = logging.getLogger(__name__)


class ExportManager:
    """مدیریت export گزارشات"""
    
    def __init__(self, db):
        self.db = db
    
    def _style_header(self, ws):
        """استایل دادن به header"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_width(self, ws):
        """تنظیم خودکار عرض ستون‌ها"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_orders(self, start_date=None, end_date=None, status=None):
        """
        Export سفارشات به Excel
        
        Args:
            start_date: تاریخ شروع (datetime)
            end_date: تاریخ پایان (datetime)
            status: فیلتر وضعیت (str)
        
        Returns:
            str: مسیر فایل ایجاد شده
        """
        try:
            # دریافت سفارشات
            conn = self.db._get_conn()
            cursor = conn.cursor()
            
            # ✅ Query ساده بدون JOIN
            query = "SELECT * FROM orders WHERE 1=1"
            params = []
            
            if start_date:
                query += " AND created_at >= ?"
                params.append(start_date.isoformat())
            
            if end_date:
                query += " AND created_at <= ?"
                params.append(end_date.isoformat())
            
            if status:
                query += " AND status = ?"
                params.append(status)
            
            query += " ORDER BY created_at DESC"
            
            cursor.execute(query, params)
            orders = cursor.fetchall()
            
            # ساخت Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "سفارشات"
            
            # Header
            headers = [
                "شماره سفارش",
                "کاربر ID",
                "محصولات",
                "قیمت کل",
                "تخفیف",
                "قیمت نهایی",
                "کد تخفیف",
                "وضعیت",
                "نحوه ارسال",
                "تاریخ ثبت"
            ]
            ws.append(headers)
            
            # داده‌ها
            for order in orders:
                try:
                    # خواندن ستون‌ها به صورت امن
                    row_data = []
                    for i, header in enumerate(headers):
                        try:
                            value = order[i] if i < len(order) else "-"
                            
                            # فرمت کردن تاریخ
                            if i == len(headers) - 1 and value and value != "-":  # ستون تاریخ
                                try:
                                    if isinstance(value, str):
                                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                        value = dt.strftime('%Y-%m-%d %H:%M')
                                except:
                                    pass
                            
                            # فرمت کردن items (JSON)
                            if i == 2 and value and value != "-":  # ستون محصولات
                                try:
                                    items = json.loads(value)
                                    value = ", ".join([f"{item.get('product', '?')} ({item.get('quantity', '?')})" for item in items])
                                except:
                                    pass
                            
                            row_data.append(value if value is not None else "-")
                        except:
                            row_data.append("-")
                    
                    ws.append(row_data)
                except Exception as e:
                    logger.error(f"Error processing order row: {e}")
                    continue
            
            # استایل
            self._style_header(ws)
            self._auto_width(ws)
            
            # ✅ ذخیره در temp directory
            temp_dir = tempfile.gettempdir()
            filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"✅ Exported {len(orders)} orders to {filename}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Error exporting orders: {e}", exc_info=True)
            raise
    
    def export_products(self):
        """Export محصولات به Excel"""
        try:
            conn = self.db._get_conn()
            cursor = conn.cursor()
            
            # ✅ Query ساده
            cursor.execute("SELECT * FROM products ORDER BY created_at DESC")
            products = cursor.fetchall()
            
            # ساخت Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "محصولات"
            
            # Header
            headers = ["ID", "نام محصول", "توضیحات", "تاریخ ایجاد"]
            ws.append(headers)
            
            # داده‌ها
            for product in products:
                try:
                    row_data = []
                    for i in range(min(4, len(product))):
                        value = product[i]
                        
                        # فرمت تاریخ
                        if i == 3 and value:
                            try:
                                if isinstance(value, str):
                                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                    value = dt.strftime('%Y-%m-%d')
                            except:
                                pass
                        
                        row_data.append(value if value is not None else "-")
                    
                    ws.append(row_data)
                except Exception as e:
                    logger.error(f"Error processing product row: {e}")
                    continue
            
            # استایل
            self._style_header(ws)
            self._auto_width(ws)
            
            # ✅ ذخیره در temp
            temp_dir = tempfile.gettempdir()
            filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"✅ Exported {len(products)} products to {filename}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Error exporting products: {e}", exc_info=True)
            raise
    
    def export_users(self):
        """Export کاربران به Excel"""
        try:
            conn = self.db._get_conn()
            cursor = conn.cursor()
            
            # ✅ Query ساده بدون JOIN
            # FIX: ستون created_at است نه joined_at
            cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
            users = cursor.fetchall()
            
            # ساخت Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "کاربران"
            
            # Header
            headers = ["User ID", "Username", "نام", "موبایل", "تاریخ عضویت"]
            ws.append(headers)
            
            # داده‌ها
            for user in users:
                try:
                    row_data = []
                    for i in range(min(5, len(user))):
                        value = user[i]
                        
                        # فرمت تاریخ
                        if i == 4 and value:
                            try:
                                if isinstance(value, str):
                                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                    value = dt.strftime('%Y-%m-%d')
                            except:
                                pass
                        
                        row_data.append(value if value is not None else "-")
                    
                    ws.append(row_data)
                except Exception as e:
                    logger.error(f"Error processing user row: {e}")
                    continue
            
            # استایل
            self._style_header(ws)
            self._auto_width(ws)
            
            # ✅ ذخیره در temp
            temp_dir = tempfile.gettempdir()
            filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"✅ Exported {len(users)} users to {filename}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Error exporting users: {e}", exc_info=True)
            raise
    
    def export_sales_report(self, period='month'):
        """
        گزارش فروش ساده
        
        Args:
            period: 'week', 'month', 'year'
        """
        try:
            # محاسبه بازه زمانی
            now = datetime.now()
            
            if period == 'week':
                start_date = now - timedelta(days=7)
                title = "هفته اخیر"
            elif period == 'month':
                start_date = now - timedelta(days=30)
                title = "ماه اخیر"
            else:  # year
                start_date = now - timedelta(days=365)
                title = "سال اخیر"
            
            conn = self.db._get_conn()
            cursor = conn.cursor()
            
            # ✅ Query ساده
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_orders,
                    SUM(final_price) as total_revenue,
                    AVG(final_price) as avg_order_value,
                    SUM(discount_amount) as total_discounts
                FROM orders
                WHERE created_at >= ? AND status IN ('confirmed', 'payment_confirmed')
            """, (start_date.isoformat(),))
            
            stats = cursor.fetchone()
            
            # ساخت Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "خلاصه"
            
            ws.append(["گزارش فروش", title])
            ws.append([])
            ws.append(["شاخص", "مقدار"])
            ws.append(["تعداد سفارشات", stats[0] or 0])
            ws.append(["جمع فروش", f"{stats[1] or 0:,} تومان"])
            ws.append(["میانگین سفارش", f"{stats[2] or 0:,.0f} تومان"])
            ws.append(["تخفیفات داده شده", f"{stats[3] or 0:,} تومان"])
            
            self._style_header(ws)
            self._auto_width(ws)
            
            # ✅ ذخیره در temp
            temp_dir = tempfile.gettempdir()
            filename = f"sales_report_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"✅ Generated sales report: {filename}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Error generating sales report: {e}", exc_info=True)
            raise


# ==================== Handler Functions ====================

async def export_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """منوی export"""
    # چک کردن ادمین
    if not update.effective_user or update.effective_user.id != ADMIN_ID:
        return
    
    keyboard = [
        [InlineKeyboardButton("📦 سفارشات", callback_data="export:orders")],
        [InlineKeyboardButton("📦 محصولات", callback_data="export:products")],
        [InlineKeyboardButton("👥 کاربران", callback_data="export:users")],
        [InlineKeyboardButton("📊 گزارش فروش (هفته)", callback_data="export:sales_week")],
        [InlineKeyboardButton("📊 گزارش فروش (ماه)", callback_data="export:sales_month")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_admin")]
    ]
    
    await update.message.reply_text(
        "📥 **دانلود گزارشات**\n\n"
        "یکی از گزینه‌های زیر را انتخاب کنید:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )


async def handle_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """پردازش export"""
    query = update.callback_query
    await query.answer()
    
    # چک کردن ادمین
    if not update.effective_user or update.effective_user.id != ADMIN_ID:
        return
    
    export_type = query.data.split(':')[1]
    
    await query.message.reply_text("⏳ در حال آماده‌سازی فایل...\n\nلطفاً کمی صبر کنید...")
    
    filepath = None
    
    try:
        db = context.bot_data.get('db')
        if not db:
            await query.message.reply_text("❌ خطا: دیتابیس در دسترس نیست!")
            return
        
        exporter = ExportManager(db)
        
        if export_type == 'orders':
            filepath = exporter.export_orders()
        elif export_type == 'products':
            filepath = exporter.export_products()
        elif export_type == 'users':
            filepath = exporter.export_users()
        elif export_type == 'sales_week':
            filepath = exporter.export_sales_report('week')
        elif export_type == 'sales_month':
            filepath = exporter.export_sales_report('month')
        else:
            await query.message.reply_text("❌ نوع export نامعتبر است!")
            return
        
        # ارسال فایل
        if filepath and os.path.exists(filepath):
            with open(filepath, 'rb') as f:
                await query.message.reply_document(
                    document=f,
                    filename=os.path.basename(filepath),
                    caption="✅ فایل آماده شد!"
                )
        else:
            await query.message.reply_text("❌ خطا در ساخت فایل!")
        
    except Exception as e:
        logger.error(f"❌ Error in export: {e}", exc_info=True)
        await query.message.reply_text(
            f"❌ خطا در ساخت فایل:\n\n"
            f"```\n{str(e)[:200]}\n```",
            parse_mode='Markdown'
        )
    
    finally:
        # ✅ FIX: حذف فایل موقت همیشه انجام میشه (حتی در صورت خطا)
        if filepath:
            try:
                os.remove(filepath)
            except Exception as cleanup_err:
                logger.warning(f"⚠️ Could not remove temp file {filepath}: {cleanup_err}")
